import re
import os
import math
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# 1. KONFIGURASI PATH
# ─────────────────────────────────────────────
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE    = os.path.join(BASE_DIR, 'input_tf-idf.txt')
STOPWORD_FILE = os.path.join(BASE_DIR, 'stopword.txt')
OUTPUT_DIR    = os.path.join(BASE_DIR, 'outputs')
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE   = os.path.join(OUTPUT_DIR, 'output_tf-idf.xlsx')

# ─────────────────────────────────────────────
# 2. LOAD STOPWORD
# ─────────────────────────────────────────────
with open(STOPWORD_FILE, 'r', encoding='utf-8') as f:
    stopwords = set(line.strip().lower() for line in f if line.strip())

# ─────────────────────────────────────────────
# 3. STEMMER
# ─────────────────────────────────────────────
factory = StemmerFactory()
stemmer = factory.create_stemmer()

# ─────────────────────────────────────────────
# 4. BACA input_tf-idf.txt
#    Format: <LABEL> = <teks dokumen>
#    Contoh: Q  = Tingkat Kurikulum 2013
#            D1 = Distribusi buku ...
# ─────────────────────────────────────────────
docs_raw   = {}   # {'Q': 'teks...', 'D1': 'teks...', ...}
doc_order  = []   # urutan label

with open(INPUT_FILE, 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if not line:
            continue
        # Pisahkan label dan teks: ambil sebelum ' = ' sebagai label
        match = re.match(r'^([A-Za-z0-9]+)\s*=\s*(.+)$', line)
        if match:
            label = match.group(1).strip().upper()
            text  = match.group(2).strip()
            docs_raw[label] = text
            doc_order.append(label)

# Pastikan urutan: Q dulu, lalu D1, D2, ...
query_label = 'Q'
doc_labels  = [l for l in doc_order if l != query_label]
all_labels  = [query_label] + doc_labels   # ['Q','D1','D2','D3','D4','D5']

N = len(doc_labels)   # jumlah dokumen (tanpa Query)

# ─────────────────────────────────────────────
# 5. PREPROCESSING (case fold + hapus non-huruf +
#    tokenisasi + hapus stopword + stemming)
# ─────────────────────────────────────────────
def preprocess(text):
    # Input sudah disederhanakan manual -- cukup lowercase + tokenisasi
    # Tidak perlu stopword removal atau stemming agar hasil sesuai referensi
    text = text.lower()
    text = re.sub(r'[^a-z0-9\s]', ' ', text)   # hilangkan tanda baca, simpan angka
    tokens = [t for t in text.split() if len(t) >= 1]
    return tokens

# Token setelah preprocessing per label
docs_tokens = {}   # {'Q': [...], 'D1': [...], ...}
for label in all_labels:
    docs_tokens[label] = preprocess(docs_raw.get(label, ''))

print("\n[INFO] Hasil Preprocessing:")
for label in all_labels:
    print(f"  {label}: {' '.join(docs_tokens[label])}")

# ─────────────────────────────────────────────
# 6. KUMPULKAN SEMUA TERM UNIK (dari dokumen D1-Dn saja, bukan Q)
#    Urutan: kemunculan pertama di seluruh dokumen
# ─────────────────────────────────────────────
seen = set()
all_terms = []
for label in all_labels:          # Q ikut berkontribusi term
    for t in docs_tokens[label]:
        if t not in seen:
            seen.add(t)
            all_terms.append(t)

all_terms.sort()   # alfabet agar rapi

# ─────────────────────────────────────────────
# 7. TF NORMALISASI: freq(term, doc) / total_terms_in_doc
# ─────────────────────────────────────────────
tf = {}   # tf[label][term] = nilai normalisasi (float)
for label in all_labels:
    tokens  = docs_tokens[label]
    total   = len(tokens)
    freq    = {}
    for t in tokens:
        freq[t] = freq.get(t, 0) + 1
    tf[label] = {
        term: (freq.get(term, 0) / total) if total > 0 else 0.0
        for term in all_terms
    }

# ─────────────────────────────────────────────
# 8. DF & IDF
#    df  = jumlah DOKUMEN (D1–Dn, bukan Q) yang mengandung term
#    IDF = log10(N / df)
# ─────────────────────────────────────────────
df  = {}
idf = {}
for term in all_terms:
    # df mencakup Q dan D1-Dn sesuai referensi
    df[term]  = sum(1 for label in all_labels if tf[label][term] > 0)
    idf[term] = math.log10(N / df[term]) if df[term] > 0 else 0.0

# ─────────────────────────────────────────────
# 9. Wdt = TF × IDF
# ─────────────────────────────────────────────
wdt = {}
for label in all_labels:
    wdt[label] = {term: tf[label][term] * idf[term]
                  for term in all_terms}

# ─────────────────────────────────────────────
# 10. TULIS KE EXCEL
# ─────────────────────────────────────────────
wb = Workbook()
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

ws = wb.create_sheet(title='TF-IDF')

# ── Helper styles ──────────────────────────
def thin():
    s = Side(style='thin')
    return Border(top=s, bottom=s, left=s, right=s)

BLUE_HDR   = PatternFill('solid', fgColor='4472C4')
BLUE_SUB   = PatternFill('solid', fgColor='9DC3E6')
YELLOW     = PatternFill('solid', fgColor='FFFF99')
GREEN      = PatternFill('solid', fgColor='E2EFDA')
WHITE      = PatternFill('solid', fgColor='FFFFFF')
GOLD_ROW   = PatternFill('solid', fgColor='FFC000')   # baris summary kuning

F_WHITE_B  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
F_BOLD     = Font(bold=True, name='Calibri', size=11)
F_NORM     = Font(name='Calibri', size=11)
F_SMALL    = Font(name='Calibri', size=10)

def hcell(ws, r, c, val, fill=None, font=None, h='center', span=1, wrap=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = font or F_WHITE_B
    cell.fill      = fill or BLUE_HDR
    cell.alignment = Alignment(horizontal=h, vertical='center', wrap_text=wrap)
    cell.border    = thin()
    return cell

def dcell(ws, r, c, val, fill=None, font=None, h='center'):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = font or F_NORM
    cell.fill      = fill or WHITE
    cell.alignment = Alignment(horizontal=h, vertical='center')
    cell.border    = thin()
    return cell

# ─────────────────────────────────────────────
# LAYOUT KOLOM
# Term | TF(Q,D1..Dn) | IDF(df, log(D/df)) | Wdt=TF×IDF(Q,D1..Dn)
#
# col 1      : Term
# col 2..    : TF per label (Q, D1, D2, ..., Dn)
# col berikut: df, log(D/df)
# col berikut: Wdt per label (Q, D1, D2, ..., Dn)
# ─────────────────────────────────────────────
COL_TERM   = 1
COL_TF_START = 2                              # Q, D1, D2, ...
COL_TF_END   = COL_TF_START + len(all_labels) - 1
COL_DF       = COL_TF_END + 1
COL_LOGD     = COL_DF + 1
COL_WDT_START = COL_LOGD + 1
COL_WDT_END   = COL_WDT_START + len(all_labels) - 1
TOTAL_COLS    = COL_WDT_END

# ── Baris 1: Judul besar ──────────────────
row = 1
c = ws.cell(row=row, column=1, value='PERHITUNGAN KE DALAM TF IDF')
c.font      = Font(bold=True, size=14, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
ws.row_dimensions[row].height = 22
row += 1

# ── Baris 2: kosong ──────────────────────
row += 1

# ── Baris 3: Info jumlah dokumen ─────────
ws.cell(row=row, column=1, value='Menentukan bobot setiap term  D =').font = F_BOLD
ws.cell(row=row, column=4, value=N).font = Font(bold=True, size=12, name='Calibri')
row += 1
row += 1   # spasi

# ── Baris 5: Header grup TF / IDF / Wdt ─
# Merge "Tf" di atas kolom TF
ws.merge_cells(start_row=row, start_column=COL_TF_START,
               end_row=row,   end_column=COL_TF_END)
hcell(ws, row, COL_TF_START, 'Tf (Normalisasi)')

# Merge "IDF" di atas kolom df + log(D/df)
ws.merge_cells(start_row=row, start_column=COL_DF,
               end_row=row,   end_column=COL_LOGD)
hcell(ws, row, COL_DF, 'IDF')

# Merge "Wdt = Tf . IDF" di atas kolom Wdt
ws.merge_cells(start_row=row, start_column=COL_WDT_START,
               end_row=row,   end_column=COL_WDT_END)
hcell(ws, row, COL_WDT_START, 'Wdt = Tf . IDF')

# isi kolom Term (merge 2 baris: row 5+6)
ws.merge_cells(start_row=row, start_column=COL_TERM,
               end_row=row+1, end_column=COL_TERM)
hcell(ws, row, COL_TERM, 'Term')
ws.row_dimensions[row].height = 18
row += 1

# ── Baris 6: Sub-header kolom ────────────
# TF: Q, D1, D2, ..., Dn
for i, label in enumerate(all_labels):
    hcell(ws, row, COL_TF_START + i, label, fill=BLUE_SUB,
          font=Font(bold=True, name='Calibri', size=11))

# IDF: df, log(D/df)
hcell(ws, row, COL_DF,   'df',         fill=BLUE_SUB,
      font=Font(bold=True, name='Calibri', size=11))
hcell(ws, row, COL_LOGD, 'log(D/df)', fill=BLUE_SUB,
      font=Font(bold=True, name='Calibri', size=11))

# Wdt: Q, D1, D2, ..., Dn
for i, label in enumerate(all_labels):
    hcell(ws, row, COL_WDT_START + i, label, fill=BLUE_SUB,
          font=Font(bold=True, name='Calibri', size=11))

ws.row_dimensions[row].height = 18
row += 1

# ── Data per term ─────────────────────────
for term in all_terms:
    # Kolom Term
    dcell(ws, row, COL_TERM, term, h='left')

    # TF (normalisasi)
    for i, label in enumerate(all_labels):
        val  = tf[label][term]
        cell = ws.cell(row=row, column=COL_TF_START + i, value=val)
        cell.font      = F_NORM
        cell.fill      = YELLOW if val > 0 else WHITE
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin()
        if val != 0:
            cell.number_format = '0.000000'

    # IDF: df
    dcell(ws, row, COL_DF,   df[term])
    # IDF: log(D/df)
    dcell(ws, row, COL_LOGD, idf[term],
          fill=GREEN if idf[term] > 0 else WHITE)

    # Wdt = TF x IDF
    for i, label in enumerate(all_labels):
        val  = wdt[label][term]
        fill = YELLOW if val > 0 else WHITE
        cell = ws.cell(row=row, column=COL_WDT_START + i, value=val)
        cell.font      = F_NORM
        cell.fill      = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin()
        if val != 0:
            cell.number_format = '0.000000'

    row += 1

# Baris "panjang docs" -- total token per dokumen
PD_FILL = PatternFill('solid', fgColor='F2F2F2')
dcell(ws, row, COL_TERM, 'panjang docs', fill=GOLD_ROW, font=F_BOLD, h='left')
for i, label in enumerate(all_labels):
    dcell(ws, row, COL_TF_START + i, len(docs_tokens[label]), fill=GOLD_ROW, font=F_BOLD)
row += 1

# ── Lebar kolom ──────────────────────────
ws.column_dimensions[get_column_letter(COL_TERM)].width = 16
for i in range(len(all_labels)):
    ws.column_dimensions[get_column_letter(COL_TF_START + i)].width  = 10
    ws.column_dimensions[get_column_letter(COL_WDT_START + i)].width = 10
ws.column_dimensions[get_column_letter(COL_DF)].width   = 5
ws.column_dimensions[get_column_letter(COL_LOGD)].width = 12

# ─────────────────────────────────────────────
# SHEET 2: Info Dokumen (Q, D1-D5)
# ─────────────────────────────────────────────
ws2 = wb.create_sheet(title='Dokumen')
row2 = 1

ws2.cell(row=row2, column=1,
         value='MENGHITUNG TF IDF DAN COSINUS SIMILARITY DENGAN KASUS BERIKUT'
         ).font = Font(bold=True, size=12, name='Calibri')
ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=10)
ws2.row_dimensions[row2].height = 20
row2 += 2

# Daftar dokumen
ws2.cell(row=row2, column=1, value=f'Misalkan terdapat {N} dokumen :').font = F_BOLD
row2 += 1
for label in doc_labels:
    c1 = ws2.cell(row=row2, column=1, value=f'{label} =')
    c1.font = F_BOLD
    c2 = ws2.cell(row=row2, column=2, value=docs_raw.get(label, ''))
    c2.font = F_NORM
    c2.alignment = Alignment(wrap_text=True)
    ws2.merge_cells(start_row=row2, start_column=2, end_row=row2, end_column=10)
    row2 += 1

row2 += 1
ws2.cell(row=row2, column=1, value='dan terdapat sebuah Query').font = F_BOLD
ws2.cell(row=row2, column=3, value='Q =').font = F_BOLD
ws2.cell(row=row2, column=4, value=docs_raw.get('Q', '')).font = F_NORM
row2 += 2

ws2.cell(row=row2, column=1,
         value='Lakukan preprosesing terhadap semua dokumen dengan tokenisasi, stemming dan penghapusan stopword'
         ).font = Font(bold=True, name='Calibri', size=11)
ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=10)
row2 += 2

# Tabel hasil preprocessing
hcell(ws2, row2, 1, 'Dokumen', fill=BLUE_HDR)
hcell(ws2, row2, 2, 'Term yang mewakili dokumen', fill=BLUE_HDR, h='left')
ws2.merge_cells(start_row=row2, start_column=2, end_row=row2, end_column=10)
row2 += 1

for label in all_labels:
    dcell(ws2, row2, 1, label)
    c = ws2.cell(row=row2, column=2, value=' '.join(docs_tokens[label]))
    c.font   = F_NORM
    c.border = thin()
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws2.merge_cells(start_row=row2, start_column=2, end_row=row2, end_column=10)
    row2 += 1

ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 80

# ─────────────────────────────────────────────
# SHEET 3: COSINE SIMILARITY
# ─────────────────────────────────────────────
ws3 = wb.create_sheet(title='COSSIM')

TEAL      = PatternFill('solid', fgColor='4F6228')   # header gelap
TEAL_SUB  = PatternFill('solid', fgColor='C4D79B')   # sub-header
GOLD_ROW  = PatternFill('solid', fgColor='FFC000')   # baris jumlah WD
BLUE_ROW  = PatternFill('solid', fgColor='538ED5')   # baris panjang vektor
BLUE_SUB2 = PatternFill('solid', fgColor='BDD7EE')

F_TEAL_B  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
F_DARK_B  = Font(bold=True, name='Calibri', size=11)

def hc3(ws, r, c, val, fill=None, font=None, h='center', wrap=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = font or F_TEAL_B
    cell.fill      = fill or TEAL
    cell.alignment = Alignment(horizontal=h, vertical='center', wrap_text=wrap)
    cell.border    = thin()
    return cell

def dc3(ws, r, c, val, fill=None, font=None, h='center', fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = font or F_NORM
    cell.fill      = fill or WHITE
    cell.alignment = Alignment(horizontal=h, vertical='center')
    cell.border    = thin()
    if fmt:
        cell.number_format = fmt
    return cell

# ── Hitung nilai WD (wq × wdi per term) ───────
# wq[term] = wdt['Q'][term]
# wd_product[Di][term] = wdt['Q'][term] * wdt['Di'][term]
# WD_sum[Di] = sum atas semua term
wd_product = {}  # wd_product[label][term]
for label in all_labels:
    wd_product[label] = {}
    for term in all_terms:
        wd_product[label][term] = wdt['Q'][term] * wdt[label][term]

WD_sum = {label: sum(wd_product[label][t] for t in all_terms)
          for label in all_labels}

# ── Hitung Panjang Vektor: pv[label] = sqrt(sum(wdt[label][t]^2)) ──
pv_sq = {}   # pv_sq[label][term] = wdt[label][term]^2
for label in all_labels:
    pv_sq[label] = {term: wdt[label][term] ** 2 for term in all_terms}

PV = {label: math.sqrt(sum(pv_sq[label][t] for t in all_terms))
      for label in all_labels}

# ── Cosine Similarity: cos(Q, Di) = WD_sum[Di] / (PV[Q] * PV[Di]) ──
cos_sim = {}
for label in doc_labels:
    denom = PV['Q'] * PV[label]
    cos_sim[label] = (WD_sum[label] / denom) if denom != 0 else 0.0

# ───────────────────────────────────────────────────────────────────
# LAYOUT KOLOM:
# Col 1         : Term
# Col 2..7      : WD table  (Q, D1..D5)
# Col 8         : spasi
# Col 9..14     : Panjang Vektor (Q, D1..D5)
# ───────────────────────────────────────────────────────────────────
C_TERM   = 1
C_WD_S   = 2                              # WD start
C_WD_E   = C_WD_S + len(all_labels) - 1  # WD end
C_GAP    = C_WD_E + 1
C_PV_S   = C_GAP + 1                     # PV start
C_PV_E   = C_PV_S + len(all_labels) - 1  # PV end

r3 = 1

# ── Baris 1: Judul ────────────────────────
c = ws3.cell(row=r3, column=1, value='PERHITUNGAN KE DALAM COSINE SIMILARITY')
c.font = Font(bold=True, size=14, name='Calibri')
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=C_PV_E)
ws3.row_dimensions[r3].height = 22
r3 += 2

# ── Baris 3: Rumus label ──────────────────
ws3.cell(row=r3, column=1, value='Rumus untuk menghitung cos similarity').font = F_BOLD
r3 += 1

# ── Baris 4: Rumus cos ────────────────────
formula_cell = ws3.cell(row=r3, column=1,
    value='cos(d, q) = Σ(wdi × wqi) / ( √Σwdi²  ×  √Σwqi² )')
formula_cell.font = Font(bold=True, italic=True, name='Calibri', size=11)
ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=C_PV_E)
r3 += 2

# ── Keterangan a & b ──────────────────────
ws3.cell(row=r3, column=1,
    value='a. Hitung perkalian skalar antara Q dan 5 Dokumen lainnya kemudian hasilnya di jumlahkan '
          '( bagian pembilang pada rumus di atas)').font = F_NORM
ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=C_PV_E)
r3 += 1
ws3.cell(row=r3, column=1,
    value='b. Hitung panjang vektor setiap dokumen dengan mengkuadratkan setiap bobot dan hasilnya '
          'di jumlahkan kemudian di akarkan (bagian penyebut pada rumus di atas)').font = F_NORM
ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=C_PV_E)
r3 += 2

# ── Header grup: WD dan Panjang Vektor ────
ws3.merge_cells(start_row=r3, start_column=C_WD_S, end_row=r3, end_column=C_WD_E)
hc3(ws3, r3, C_WD_S, 'WD = WQ . Wdt')

ws3.merge_cells(start_row=r3, start_column=C_PV_S, end_row=r3, end_column=C_PV_E)
hc3(ws3, r3, C_PV_S, 'Panjang Vektor')

# Merge term (2 baris)
ws3.merge_cells(start_row=r3, start_column=C_TERM, end_row=r3+1, end_column=C_TERM)
hc3(ws3, r3, C_TERM, 'Term')
ws3.row_dimensions[r3].height = 18
r3 += 1

# ── Sub-header kolom ──────────────────────
for i, label in enumerate(all_labels):
    hc3(ws3, r3, C_WD_S + i, label, fill=TEAL_SUB,
        font=Font(bold=True, name='Calibri', size=11, color='000000'))
    hc3(ws3, r3, C_PV_S + i, label, fill=BLUE_SUB2,
        font=Font(bold=True, name='Calibri', size=11, color='000000'))

ws3.row_dimensions[r3].height = 18
r3 += 1

# ── Data per term ─────────────────────────
for term in all_terms:
    dc3(ws3, r3, C_TERM, term, h='left')
    for i, label in enumerate(all_labels):
        val_wd = wd_product[label][term]
        fill_wd = YELLOW if val_wd > 0 else WHITE
        cell_wd = dc3(ws3, r3, C_WD_S + i, val_wd, fill=fill_wd)
        if val_wd != 0:
            cell_wd.number_format = '0.000000'

        val_pv = pv_sq[label][term]
        fill_pv = GREEN if val_pv > 0 else WHITE
        cell_pv = dc3(ws3, r3, C_PV_S + i, val_pv, fill=fill_pv)
        if val_pv != 0:
            cell_pv.number_format = '0.000000'
    r3 += 1

# ── Baris Jumlah WD (kuning) ──────────────
dc3(ws3, r3, C_TERM, 'Jumlah', fill=GOLD_ROW, font=F_BOLD)
for i, label in enumerate(all_labels):
    dc3(ws3, r3, C_WD_S + i, WD_sum[label], fill=GOLD_ROW, font=F_BOLD, fmt='0.000000')
ws3.row_dimensions[r3].height = 16
r3 += 1

# ── Baris Panjang Vektor (biru) ───────────
dc3(ws3, r3, C_TERM, 'Panjang Vektor', fill=BLUE_ROW,
    font=Font(bold=True, color='FFFFFF', name='Calibri', size=11))
for i, label in enumerate(all_labels):
    dc3(ws3, r3, C_PV_S + i, PV[label], fill=BLUE_ROW,
        font=Font(bold=True, color='FFFFFF', name='Calibri', size=11), fmt='0.000000')
ws3.row_dimensions[r3].height = 16
r3 += 3

# ── Cos Similarity hasil ──────────────────
ws3.cell(row=r3, column=1,
    value='Menghitung nilai Cosine similarity masing - masing dokumen dengan query'
    ).font = Font(bold=True, size=11, name='Calibri')
ws3.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=C_PV_E)
r3 += 1

# Cos(Q,Di) kiri — Tabel ranking kanan
cos_col_start = 1
rank_col_start = 5

# Simpan baris awal untuk tabel ranking (sama dengan baris pertama cos list)
rank_row_start = r3

for label in doc_labels:
    ws3.cell(row=r3, column=cos_col_start,
             value=f'Cos (Q,{label}) =').font = F_BOLD
    cell_cos = ws3.cell(row=r3, column=cos_col_start + 2, value=cos_sim[label])
    cell_cos.font          = F_NORM
    cell_cos.number_format = '0.000000'
    r3 += 1

# Tabel ranking di sebelah kanan mulai dari rank_row_start
r_rank = rank_row_start

# Label "Hasil perhitungan di:" dan "Jika diurutkan"
ws3.cell(row=r_rank, column=rank_col_start,     value='Hasil perhitungan di:').font = F_BOLD
ws3.cell(row=r_rank, column=rank_col_start + 3, value='Jika diurutkan').font = F_BOLD
r_rank += 1

# Header kolom tabel kiri (urutan asli D1-D5)
hc3(ws3, r_rank, rank_col_start,     'Dokumen', fill=TEAL)
hc3(ws3, r_rank, rank_col_start + 1, 'Hasil',   fill=TEAL)

# Header kolom tabel kanan (diurutkan)
hc3(ws3, r_rank, rank_col_start + 3, 'Dokumen', fill=BLUE_HDR)
hc3(ws3, r_rank, rank_col_start + 4, 'Hasil',   fill=BLUE_HDR)
r_rank += 1

# Isi tabel kiri (urutan asli)
r_rank_left = r_rank
for label in doc_labels:
    dc3(ws3, r_rank_left, rank_col_start,     label,          fill=TEAL_SUB)
    dc3(ws3, r_rank_left, rank_col_start + 1, cos_sim[label], fmt='0.000000')
    r_rank_left += 1

# Isi tabel kanan (diurutkan descending)
sorted_docs = sorted(cos_sim.items(), key=lambda x: x[1], reverse=True)
r_rank_right = r_rank
for label, val in sorted_docs:
    dc3(ws3, r_rank_right, rank_col_start + 3, label, fill=BLUE_SUB2)
    dc3(ws3, r_rank_right, rank_col_start + 4, val,   fmt='0.000000')
    r_rank_right += 1

# ── Kesimpulan ────────────────────────────
best_doc, best_val = sorted_docs[0]
kc_row = max(r3, r_rank_left, r_rank_right) + 2
ws3.cell(row=kc_row, column=1, value='Kesimpulan').font = Font(bold=True, size=12, name='Calibri')
kc_row += 1
conc_text = (
    f'Berdasarkan perhitungan Cosine Similarity, dokumen yang paling relevan dengan query adalah '
    f'{best_doc} dengan nilai cos = {best_val:.6f}. '
    f'Urutan relevansi dokumen dari tertinggi ke terendah: '
    + ', '.join(f'{lbl} ({val:.6f})' for lbl, val in sorted_docs) + '.'
)
c = ws3.cell(row=kc_row, column=1, value=conc_text)
c.font = F_NORM
c.alignment = Alignment(wrap_text=True)
ws3.merge_cells(start_row=kc_row, start_column=1, end_row=kc_row + 2, end_column=C_PV_E)
ws3.row_dimensions[kc_row].height = 40

# ── Lebar kolom ───────────────────────────
ws3.column_dimensions[get_column_letter(C_TERM)].width = 16
for i in range(len(all_labels)):
    ws3.column_dimensions[get_column_letter(C_WD_S + i)].width = 9
    ws3.column_dimensions[get_column_letter(C_PV_S + i)].width = 9
ws3.column_dimensions[get_column_letter(C_GAP)].width = 2

# ─────────────────────────────────────────────
# SIMPAN
# ─────────────────────────────────────────────
wb.save(OUTPUT_FILE)

print(f"\n[OK]  Output berhasil disimpan : {OUTPUT_FILE}")
print(f"[DOC] Jumlah dokumen (D)       : {N}")
print(f"[TRM] Jumlah term unik         : {len(all_terms)}")
print(f"[SHT] Sheet                    : TF-IDF | Dokumen | COSSIM")
