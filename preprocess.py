import pandas as pd
import re
import os
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
from collections import Counter

# ─────────────────────────────────────────────
# 1. KONFIGURASI PATH (relatif ke lokasi script)
# ─────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE   = os.path.join(BASE_DIR, 'input.txt')
STOPWORD_FILE= os.path.join(BASE_DIR, 'stopword.txt')
OUTPUT_DIR   = os.path.join(BASE_DIR, 'outputs')
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE  = os.path.join(OUTPUT_DIR, 'output_preprocessing.xlsx')

# ─────────────────────────────────────────────
# 2. LOAD STOPWORD
# ─────────────────────────────────────────────
with open(STOPWORD_FILE, 'r', encoding='utf-8') as f:
    stopwords = set(line.strip().lower() for line in f if line.strip())

# ─────────────────────────────────────────────
# 3. STEMMER
# ─────────────────────────────────────────────
factory = StemmerFactory()
stemmer = factory.create_stemmer()

# ─────────────────────────────────────────────
# 4. BACA INPUT.TXT
#    Baris 1 = Judul, Baris 2 = Sumber, sisa = isi artikel
# ─────────────────────────────────────────────
with open(INPUT_FILE, 'r', encoding='utf-8') as f:
    raw_lines = [line.rstrip('\n') for line in f]

# Pisahkan header (judul + sumber) dari isi
judul  = raw_lines[0].strip() if len(raw_lines) > 0 else ''
sumber = raw_lines[1].strip() if len(raw_lines) > 1 else ''
isi_lines = raw_lines[2:]
isi_text = ' '.join(line for line in isi_lines if line.strip())

# ─────────────────────────────────────────────
# 5. PREPROCESSING FULL TEXT
#    Hasilkan tabel: setiap token unik dengan frekuensi & keterangan
# ─────────────────────────────────────────────

# 5a. Ekstrak angka dari teks ASLI beserta konteks (karakter sebelumnya)
def label_angka(num_str, before=''):
    """Beri label kontekstual pada angka."""
    # Uang: didahului Rp atau US (dari US$)
    if re.search(r'(?:Rp|US)$', before.strip()):
        return f'{num_str}(uang)'
    # Waktu: pola HH.MM atau HH:MM (jam 0-23, menit 0-59)
    m = re.match(r'^(\d{1,2})[.:](\d{2})$', num_str)
    if m and int(m.group(1)) <= 23 and int(m.group(2)) <= 59:
        return f'{num_str}(waktu)'
    # Tahun: 4 digit range 1900-2099
    if re.match(r'^\d{4}$', num_str) and 1900 <= int(num_str) <= 2099:
        return f'{num_str}(tahun)'
    return num_str

# Cari semua angka + 4 karakter sebelumnya untuk konteks
raw_number_matches = list(re.finditer(r'\d+(?:[.,]\d+)*', isi_text))
number_labels = []
for m in raw_number_matches:
    before = isi_text[max(0, m.start()-4):m.start()]  # konteks sebelum angka
    number_labels.append(label_angka(m.group(), before))
raw_numbers    = [m.group() for m in raw_number_matches]
num_frekuensi  = len(raw_numbers)
num_keterangan = ', '.join(number_labels) if number_labels else ''

# 5b. Case folding & bersihkan angka/tanda baca untuk kata
clean_text = isi_text.lower()
clean_text = re.sub(r'[^a-z\s]', ' ', clean_text)
all_tokens_raw = clean_text.split()

# Abaikan token kosong dan sangat pendek (≤1 huruf)
all_tokens_raw = [t for t in all_tokens_raw if len(t) > 1]

# 5b. Hitung frekuensi kata MENTAH (sebelum preprocessing)
raw_freq = Counter(all_tokens_raw)

# 5c. Helper: deteksi imbuhan yang dihapus
def get_affix_label(word, stem):
    """Kembalikan label imbuhan yg dilepas, misal '(-kan)', '(me-)', '(me-, -kan)'."""
    if stem == word:
        return ''
    idx = word.find(stem)
    if idx >= 0:
        prefix = word[:idx]          # imbuhan depan
        suffix = word[idx+len(stem):]  # imbuhan belakang
        parts = []
        if prefix:
            parts.append(f'{prefix}-')
        if suffix:
            parts.append(f'-{suffix}')
        if parts:
            return 'Stemming (' + ', '.join(parts) + ')'
    # Fallback: stem tidak tampil literal (perubahan nasal, dll)
    return f'Stemming ({word}\u2192{stem})'

# 5d. Bangun tabel preprocessing per kata unik
rows = []

for word, freq in sorted(raw_freq.items(), key=lambda x: -x[1]):
    is_stopword = word in stopwords
    stem = stemmer.stem(word)

    if is_stopword:
        keterangan = 'Stopword'
    elif stem != word:
        keterangan = get_affix_label(word, stem)
    else:
        keterangan = ''   # kata biasa, tidak berubah

    rows.append({
        'word_raw'   : word,
        'stem'       : stem if not is_stopword else word,
        'freq'       : freq,
        'is_stop'    : is_stopword,
        'keterangan' : keterangan
    })

# 5e. Nomor baris (No.) sesuai urutan kemunculan pertama di teks
order_map = {}
idx = 1
for t in all_tokens_raw:
    if t not in order_map:
        order_map[t] = idx
        idx += 1

rows_sorted = sorted(rows, key=lambda r: -r['freq'])   # sort by freq desc

# Sisipkan baris ANGKA di posisi yang tepat berdasarkan frekuensi
if num_frekuensi > 0:
    # Cari posisi insert agar urutan frekuensi tetap terjaga
    insert_pos = len(rows_sorted)  # default: paling bawah
    for i, r in enumerate(rows_sorted):
        if r['freq'] <= num_frekuensi:
            insert_pos = i
            break
    rows_sorted.insert(insert_pos, {
        'word_raw'   : 'angka',
        'stem'       : 'Angka',
        'freq'       : num_frekuensi,
        'is_stop'    : False,
        'keterangan' : num_keterangan
    })
    order_map['angka'] = idx  # taruh di akhir nomor urut

# ─────────────────────────────────────────────
# 6. KATA KUNCI = token NON-stopword dengan frekuensi tertinggi
# ─────────────────────────────────────────────
# Gunakan frekuensi stem (gabungkan semua kata yg punya stem sama)
stem_freq = Counter()
for r in rows:
    if not r['is_stop']:
        stem_freq[r['stem']] += r['freq']

# Ambil top-N (misal top-7)
TOP_N = 7
keywords = [word for word, _ in stem_freq.most_common(TOP_N)]

# ─────────────────────────────────────────────
# 7. TULIS KE EXCEL
# ─────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    wb = writer.book

    # ── Buat worksheet manual ──────────────────
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title='Berita1')

    # Hapus sheet default jika ada
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']

    # ── Helper styles ──────────────────────────
    def thin_border(top=True, bottom=True, left=True, right=True):
        s = Side(style='thin')
        return Border(
            top    = s if top    else None,
            bottom = s if bottom else None,
            left   = s if left   else None,
            right  = s if right  else None,
        )

    BLUE_FILL   = PatternFill('solid', fgColor='4472C4')   # header tabel
    LIGHT_FILL  = PatternFill('solid', fgColor='BDD7EE')   # keterangan berwarna
    HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    BOLD_FONT   = Font(bold=True, name='Calibri', size=11)
    NORM_FONT   = Font(name='Calibri', size=11)
    RED_FONT    = Font(color='FF0000', name='Calibri', size=11, underline='single')
    BLUE_FONT   = Font(color='0070C0', name='Calibri', size=11, underline='single')

    row = 1  # current Excel row pointer

    # ── Judul artikel ──────────────────────────
    ws.cell(row=row, column=1, value=judul).font = Font(bold=True, size=12, name='Calibri')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    # ── Sumber (biru, underline) ───────────────
    c = ws.cell(row=row, column=1, value=sumber)
    c.font = BLUE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    row += 1   # spasi

    # ── Isi artikel ───────────────────────────
    for line in isi_lines:
        if not line.strip():
            row += 1
            continue
        c = ws.cell(row=row, column=1, value=line.strip())
        c.font = NORM_FONT
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    row += 1   # spasi sebelum KATA KUNCI

    # ── KATA KUNCI ─────────────────────────────
    ws.cell(row=row, column=1, value='KATA KUNCI').font = BOLD_FONT
    row += 1
    for kw in keywords:
        c = ws.cell(row=row, column=1, value=kw)
        c.font = BLUE_FONT   # biru seperti di foto
        row += 1

    row += 1   # spasi

    # ── HASIL PREPROCESSING ────────────────────
    ws.cell(row=row, column=1, value='HASIL PREPROCESSING').font = BOLD_FONT
    row += 1

    # Header tabel
    headers = ['No.', 'Kata', 'Frekuensi', 'Keterangan']
    col_widths = [6, 18, 12, 40]
    for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=row, column=col_i, value=h)
        c.font      = HEADER_FONT
        c.fill      = BLUE_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_i)].width = w
    row += 1

    # Data tabel
    for r_data in rows_sorted:
        no_val = order_map.get(r_data['word_raw'], '')
        # Tampilkan kata DASAR (stem) di kolom Kata
        kata   = r_data['stem']
        freq_v = r_data['freq']
        ket    = r_data['keterangan']

        vals = [no_val, kata, freq_v, ket]
        for col_i, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col_i, value=val)
            c.font      = NORM_FONT
            c.border    = thin_border()
            c.alignment = Alignment(horizontal='center' if col_i in (1, 3) else 'left',
                                    vertical='center')
            # Keterangan berwarna jika stopword / stemming
            if col_i == 4 and ket:
                c.fill = LIGHT_FILL
                c.font = Font(color='C00000', name='Calibri', size=11)
        row += 1

    # Set lebar kolom A lebih lebar untuk teks berita
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['E'].width = 5

print(f"✅  Output berhasil disimpan: {OUTPUT_FILE}")
print(f"📝  Kata Kunci: {', '.join(keywords)}")
print(f"📊  Total kata unik: {len(rows)}")
